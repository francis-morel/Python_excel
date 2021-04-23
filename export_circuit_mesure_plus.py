import openpyxl
from os import listdir


class CircuitCount:  # définition d'une classe pour mettre plusieurs informations dans un dictionnaire selon le nom du circuit
    circuit = ""
    mesure = 0
    nombre = 0

    def __init__(self, circuit, mesure):
        self.circuit = circuit
        self.mesure = mesure

    def printValues(self):
        print(self.circuit, self.mesure, "mm x", self.nombre)



class ExcelWriter:

    def __init__(self):
        self.file = openpyxl.Workbook()
        self.sheet = self.file.active
        self.fileRow = 1

    def save(self, filename):
        self.file.save(filename)

    def getRow(self):
        return self.fileRow

    def setRow(self, row):
        self.fileRow = row

    def nextRow(self):
        self.fileRow += 1

    def skipRow(self, nb_rows):
        self.fileRow += nb_rows

    def valid_sheet_name(name):
        return len(name) == 1

    def write_sheet(self, data, col=1, row=0):
        if row == 0:
            row = self.fileRow
        self.sheet.cell(row=row, column=col).value = data
        # self.nextRow()

    def write_sheet_list(self, data, start_col=1, start_row=0):
        if start_row == 0:
            start_row = self.fileRow
        for i in range(len(data)):
            self.write_sheet(data[i], start_col, start_row)
            start_row += 1

    def write_sheet_2d_list(self, data, start_col=1, start_row=1):
        pass


feuille1 = ExcelWriter()
feuille1.write_sheet_list(["allo", "salut", "bonjour", "blergh"])
feuille1.write_sheet_list(["allo", "salut", "bonjour", "blergh"], start_col=3)
feuille1.save("coucou.xlsx")


def valid_sheet_name(name):
    return len(name) == 1


def get_column(sheet, col, start_row=1, nb_lines=0):
    array = []
    row_number = start_row

    while True:
        if nb_lines > 0 and row_number-start_row > nb_lines:
            break
        cell = get_cell(sheet, col, row_number)
        if cell == None:
            break
        array.append(cell)

        row_number += 1

    return array


def get_cell(sheet, col, row):
    return sheet.cell(row=row, column=col).value


def write_sheet_list(sheet, data, start_col=1, start_row=1):
    for i in range(len(data)):
        write_sheet(sheet, data[i], start_col, start_row+i)


def write_sheet(sheet, data, col, row):
    sheet.cell(row=row, column=col).value = data


def ask_header():
    print("Veuillez entrer l'en-tête et écrire done lorsque terminé")
    en_tete = []
    while True:
        entre = input()
        if entre == "done" or entre == "Done":
            break
        en_tete.append(entre)
    return en_tete


newWb = openpyxl.Workbook()
newWs = newWb.active
fileRow = 1

totalCables = {}


for file in listdir():  # Itérer tous les fichiers dans le dossier
    if file[-5:] == ".xlsm" or file[-5:] == ".xlsx":  # Si le fichier est un fichier excel
        # Écrire le nom du fichier dans le nouveau fichier
        write_sheet(newWs, file, 1, fileRow)
        fileRow += 1
        # Ouvrir le fichier excel
        wb = openpyxl.load_workbook(filename=file, data_only=True)

        writeCol = 1
        for sheet in wb.worksheets:  # Itérer toutes les feuilles du fichier ouvert

            # Vérification du nom de la feuille et la skipper si elle ne m'intéresse pas
            if not valid_sheet_name(sheet.title):
                continue

            ColCircuit = get_column(sheet, 4)
            ColMesure = get_column(sheet, 5)

            circuitRow = fileRow

            write_sheet_list(newWs, ColCircuit,
                             start_col=writeCol, start_row=fileRow)
            writeCol += 1
            write_sheet_list(newWs, ColMesure,
                             start_col=writeCol, start_row=fileRow)
            writeCol += 2

        fileRow += len(ColCircuit) + 1

        # Fin du fichier, préparation des variables pour le prochain.
        fileRow += 2


newWb.save("total modeles python plus.xlsx")
