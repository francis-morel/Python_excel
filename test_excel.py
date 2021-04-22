import openpyxl
from os import listdir

class CircuitCount: #définition d'une classe pour mettre plusieurs informations dans un dictionnaire selon le nom du circuit
    circuit = ""
    mesure = 0
    nombre = 1

    def __init__(self, circuit, mesure):
        self.circuit = circuit
        self.mesure = mesure

    def printValues(self):
        print(self.circuit, self.mesure, "mm x", self.nombre)


newWb = openpyxl.Workbook()
newWs = newWb.active
writeCol = 1
returnColCircuit = []
returnColMesure = []
fileRow = 1

totalCables = {}

for file in listdir(): #Itérer tous les fichiers dans le dossier
    if file[-5:] == ".xlsm" or file[-5:] == ".xlsx": #Si le fichier est un fichier excel
        newWs.cell(row = fileRow, column = 1).value = file #Écrire le nom du fichier dans le nouveau fichier
        fileRow += 1
        wb = openpyxl.load_workbook(filename = file, data_only = True) #Ouvrir le fichier excel

        for sheet in wb.worksheets: #Itérer toutes les feuilles du fichier ouvert
                        
            if len(sheet.title) > 1: #Vérification du nom de la feuille et la skipper si elle ne m'intéresse pas
                continue

            for i in range(1,100): #Lire les colonnes 4 et 5 de la feuille
                cell = sheet.cell(row = i+1, column = 4)
                cell2 = sheet.cell(row = i+1, column = 5)
                nombrePanneaux = sheet.cell(row = 31, column = 3) #Lire une cellule en particulier pour obtenir le nombre de répétitions de cette entrée
                if cell.value == None:
                    break
                returnColCircuit.append(cell.value)
                returnColMesure.append(cell2.value)
                
                #Ajout des informations dans un dictionnaire pour faire un total à la fin
                if cell.value not in totalCables.keys():
                    totalCables[cell.value] = CircuitCount(cell.value, cell2.value) #Ajout de la key dans mon dictionnaire des totaux
                else:
                    totalCables[cell.value].nombre += int(nombrePanneaux.value) #Ajout du nombre de fois que ce type se répète

            circuitRow = fileRow
            for i in range(len(returnColCircuit)): #Écrire les colonnes que je viens de lire dans le nouveau fichier
                newWs.cell(row = circuitRow, column = writeCol).value = returnColCircuit[i]
                newWs.cell(row = circuitRow, column = writeCol+1).value = returnColMesure[i]
                circuitRow += 1

            returnColCircuit = []
            returnColMesure = []

            writeCol += 3

        #Fin du fichier, préparation des variables pour le prochain.
        fileRow = circuitRow + 1
        writeCol = 1

newWs.cell(row = fileRow, column = 1).value = "Total ultime:"   
fileRow += 1
for key in totalCables.keys():
    newWs.cell(row = fileRow, column = 1).value = totalCables[key].circuit
    newWs.cell(row = fileRow, column = 2).value = totalCables[key].mesure
    newWs.cell(row = fileRow, column = 3).value = totalCables[key].nombre
    fileRow += 1


newWb.save("test.xlsx")