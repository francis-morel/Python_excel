import openpyxl

from os import listdir


def main():

    newWb = openpyxl.Workbook()
    newWs = newWb.active
    fileRow = 1
    totalCables = {}

    for file in listdir():  # Itérer tous les fichiers dans le dossier

        if file[-5:] == ".xlsm" or file[-5:] == ".xlsx":  # Si le fichier est un fichier excel
            # Écrire le nom du fichier dans le nouveau fichier
            write_sheet(newWs, file, 1, fileRow)
            fileRow += 1
            # Ouvrir le fichier excel
            wb = openpyxl.load_workbook(filename=file, data_only=True)
            writeCol = 1
            quantite_max_colonne = 0

            for sheet in wb.worksheets:  # Itérer toutes les feuilles du fichier ouvert
                # Vérification du nom de la feuille et la skipper si elle ne m'intéresse pas

                if not valid_sheet_name(sheet.title):
                    continue

                colCircuit = get_column(sheet, 4) #Lire la colonne 4 (circuit)
                colMesure = get_column(sheet, 5) #Lire la colonne 5 (mesure)
                nb_panneaux = get_cell(sheet, 3, 31) #Lire la case nombre de panneaux (C31)
                total_dictionnaire(totalCables, colCircuit[1:], colMesure[1:], nb_panneaux) #Appel de la fonction pour faire le total dans le dictionnaire

                #circuitRow = fileRow #plus utilisé
                write_sheet_list(newWs, colCircuit, start_col=writeCol, start_row=fileRow)
                writeCol += 1 #Tasser de une colonne vers la droite pour écrire la mesure
                write_sheet_list(newWs, colMesure, start_col=writeCol, start_row=fileRow)
                writeCol += 2 #Tasser de 2 colonnes vers la droite pour écrire les circuits de la prochaine feuille

                if len(colCircuit) > quantite_max_colonne:#Déterminer quelle colonne du fichier était la plus longue
                    quantite_max_colonne = len(colCircuit)

            fileRow += quantite_max_colonne + 1 #Descendre le "curseur" de la quantité de données que l'on vient d'ajouter
            # Fin du fichier, préparation des variables pour le prochain.
            fileRow += 2

    write_sheet(newWs, "Total Spécial:", 1, fileRow) #Écriture de texte fixe dans le fichier avant d'écrire le total des circuits
    fileRow += 1
    write_sheet(newWs, "Circuit", 1, fileRow)
    write_sheet(newWs, "Mesure", 2, fileRow)
    write_sheet(newWs, "Nombre", 3, fileRow)
    fileRow += 1

    for key in totalCables.keys(): #Boucle qui fait le tour du dictionnaire pour écrire les totaux
        write_sheet(newWs, totalCables[key].circuit, 1, fileRow)
        write_sheet(newWs, totalCables[key].mesure, 2, fileRow)
        write_sheet(newWs, totalCables[key].nombre, 3, fileRow)
        fileRow += 1

        print(totalCables[key].circuit, totalCables[key].mesure, totalCables[key].nombre) #Affichier les totaux à la console pour rendre ça fancy

    newWb.save("total modeles python plus.xlsx")

    #Mettre le logiciel en pause pour le côté fancy
    print()
    input("Press Enter to exit")



class CircuitCount:  # définition d'une classe pour mettre plusieurs informations dans un dictionnaire selon le nom du circuit
    circuit = ""
    mesure = 0
    nombre = 0

    def __init__(self, circuit, mesure):
        self.circuit = circuit
        self.mesure = mesure

    def print_values(self):
        print(self.circuit, self.mesure, self.nombre)

    def set_circuit(self, circuit):
        self.circuit = circuit

    def set_mesure(self, mesure):
        self.mesure = mesure

    def set_nombre(self, nombre):
        self.nombre = nombre

    def get_circuit(self):
        return self.circuit

    def get_mesure(self):
        return self.mesure

    def get_nombre(self):
        return self.nombre


class ExcelWriter:
    def __init__(self):
        self.file = openpyxl.Workbook()
        self.sheet = self.file.active
        self.file_row = 1

    def save(self, filename):
        self.file.save(filename)

    def get_row(self):
        return self.file_row

    def set_row(self, row):
        self.file_row = row

    def next_row(self):
        self.file_row += 1

    def skip_row(self, nb_rows):
        self.file_row += nb_rows

    def valid_sheet_name(name):
        return len(name) == 1

    def write_sheet(self, data, col=1, row=0):
        if row == 0:
            row = self.file_row
        self.sheet.cell(row=row, column=col).value = data
        # self.nextRow()

    def write_sheet_list(self, data, start_col=1, start_row=0):
        if start_row == 0:
            start_row = self.file_row

        for i in range(len(data)):
            self.write_sheet(data[i], start_col, start_row)
            start_row += 1

    def write_sheet_2d_list(self, data, start_col=1, start_row=1):
        pass

def valid_sheet_name(name):
    return len(name) == 1


def get_column(sheet, col, start_row=1, nb_lines=0):
    array = []
    row_number = start_row
    while True:
        if nb_lines > 0 and row_number-start_row > nb_lines:
            break

        cell = get_cell(sheet, col, row_number)
        if cell == None:
            break

        array.append(cell)
        row_number += 1

    return array


def get_cell(sheet, col, row):
    return sheet.cell(row=row, column=col).value


def write_sheet_list(sheet, data, start_col=1, start_row=1):
    for i in range(len(data)):
        write_sheet(sheet, data[i], start_col, start_row+i)

def write_sheet(sheet, data, col, row):
    sheet.cell(row=row, column=col).value = data


def ask_header():
    print("Veuillez entrer l'en-tête et écrire done lorsque terminé")
    en_tete = []
    while True:
        entre = input()
        if entre == "done" or entre == "Done":
            break

        en_tete.append(entre)

    return en_tete

def total_dictionnaire(dict, arrayCircuit, arrayMesure, addition):
    for i in range(len(arrayCircuit)):
        if arrayCircuit[i] not in dict.keys(): #si la key n'existe pas dans le dictionnaire, l'ajouter
            dict[arrayCircuit[i]] = CircuitCount(arrayCircuit[i], arrayMesure[i])
            dict[arrayCircuit[i]].nombre = addition
        else: #si la key existe déjà, incrémenter la valeur de cette key
            dict[arrayCircuit[i]].nombre += addition


if __name__ == "__main__":
    main()
